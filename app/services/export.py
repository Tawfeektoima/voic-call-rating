"""
Intelligence Export Engine — Data Science Ready Dataset (Task 64)

Produces a ~50-column analytical .xlsx from a multi-table join across
calls, employees, campaigns, call_outcomes, and agent_mastery_stats.

Features:
- JSON flattening (campaign_specific_data, strengths, weaknesses, emotion_timeline)
- Duration formatting (MM:SS + raw seconds)
- Acoustic emotion percentage calculation
- Temporal feature extraction (call_hour, call_day)
- NaN handling for legacy calls
- Styled .xlsx output with formatted column headers
"""

import io
import json
from collections import Counter
from datetime import datetime
from typing import Optional, List

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session, joinedload

from app.models import (
    Call, Employee, Campaign, CallOutcome,
    AgentMasteryStats, CallStatus, UserRole, EmployeeTeamAssignment,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]


def _seconds_to_mmss(seconds: Optional[float]) -> str:
    """Convert seconds to MM:SS string.  Returns '' for None/NaN."""
    if seconds is None or pd.isna(seconds):
        return ""
    seconds = max(0, float(seconds))
    m, s = divmod(int(seconds), 60)
    return f"{m:02d}:{s:02d}"


def _safe_json(value) -> dict:
    """Return a dict from a JSON column value (handles None, str, dict)."""
    if value is None:
        return {}
    if isinstance(value, str):
        try:
            return json.loads(value)
        except (json.JSONDecodeError, TypeError):
            return {}
    if isinstance(value, dict):
        return value
    return {}


def _safe_list(value) -> list:
    """Return a list from a JSON column value."""
    if value is None:
        return []
    if isinstance(value, str):
        try:
            return json.loads(value)
        except (json.JSONDecodeError, TypeError):
            return []
    if isinstance(value, list):
        return value
    return []


def _parse_strengths(strengths_raw) -> str:
    """Convert strengths list to a readable semicolon-separated string."""
    items = _safe_list(strengths_raw)
    if not items:
        return ""
    return "; ".join(str(s) for s in items[:5])


def _parse_top_weaknesses(weaknesses_raw, top_n: int = 3):
    """Return (text_block, top_1_issue, top_2_issue, top_3_issue)."""
    items = _safe_list(weaknesses_raw)
    issues = []
    text_parts = []
    for w in items:
        if isinstance(w, dict):
            issue = w.get("issue", "")
            detail = w.get("detail", "")
            deduction = w.get("deduction", 0)
            issues.append(issue)
            text_parts.append(f"{issue} (-{deduction}): {detail}")
        elif isinstance(w, str):
            issues.append(w)
            text_parts.append(w)

    text_block = "; ".join(text_parts) if text_parts else ""

    top = (issues + ["", "", ""])[:top_n]
    return text_block, top[0], top[1], top[2]


def _calc_emotion_percentages(emotion_timeline_raw) -> dict:
    """
    Calculate percentage of each emotion from the emotion_timeline JSON.
    Returns dict with keys like pct_calm, pct_stress, pct_agitation, etc.
    """
    items = _safe_list(emotion_timeline_raw)
    if not items:
        return {
            "pct_calm": None, "pct_stress": None, "pct_agitation": None,
            "pct_neutral": None, "pct_happy": None, "pct_sad": None,
            "pct_angry": None,
        }

    counter = Counter()
    for entry in items:
        if isinstance(entry, dict):
            emotion = str(entry.get("emotion", "neutral")).lower().strip()
            counter[emotion] += 1

    total = sum(counter.values()) or 1

    return {
        "pct_calm": round((counter.get("calm", 0) / total) * 100, 1),
        "pct_stress": round((counter.get("stress", 0) / total) * 100, 1),
        "pct_agitation": round((counter.get("agitation", 0) / total) * 100, 1),
        "pct_neutral": round((counter.get("neutral", 0) / total) * 100, 1),
        "pct_happy": round((counter.get("happy", 0) / total) * 100, 1),
        "pct_sad": round((counter.get("sad", 0) / total) * 100, 1),
        "pct_angry": round((counter.get("angry", 0) / total) * 100, 1),
    }


# ---------------------------------------------------------------------------
# Core Export Service
# ---------------------------------------------------------------------------

class ExportService:
    """Builds a Data-Science-Ready analytical dataset from the platform."""

    # Column order for the final output (Task 65 - 50 Column Engine)
    COLUMN_ORDER = [
        # --- Group 1: Identity ---
        "call_id", "call_date", "call_hour", "call_day", "status",
        # --- Group 2: Agent ---
        "agent_id", "agent_name", "agent_code", "agent_department", "agent_tier", "agent_tenure_days",
        # --- Group 3: Campaign ---
        "campaign_id", "campaign_name", "campaign_type", "campaign_status",
        # --- Group 4: Mastery ---
        "mastery_rapport", "mastery_sync", "mastery_trust", "mastery_clarity",
        # --- Group 5: Evaluation ---
        "qa_score", "overridden_score", "final_score", "is_golden_moment", "lead_status",
        "opening_ok", "closing_ok", "dob_verified",
        # --- Group 6: Talk ---
        "audio_duration_sec", "audio_duration_fmt", "agent_talk_time_sec", "agent_talk_time_fmt",
        "customer_talk_time_sec", "customer_talk_time_fmt", "talk_ratio_pct",
        # --- Group 7: Acoustic ---
        "pct_calm", "pct_stress", "pct_agitation", "pct_neutral", "pct_happy", "pct_sad", "pct_angry",
        # --- Group 8: Emotion Delta ---
        "emotion_start", "emotion_end", "de_escalation_success",
        # --- Group 9: NLP ---
        "filler_words_count", "interruptions_count", "avg_response_time_sec", "calls_before_this",
        # --- Group 10: Business Outcome ---
        "primary_outcome", "outcome_value", "follow_up_required",
        "issue_resolved", "resolution_type", "escalated", "customer_satisfaction", "repeat_call",
    ]

    @classmethod
    def build_dataset(
        cls,
        db: Session,
        campaign_id: Optional[int] = None,
        department: Optional[str] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        agent_role: Optional[UserRole] = None,
        current_user_role: Optional[UserRole] = None,
        qa_scope_team_id: Optional[int] = None,
        qa_scope_campaign_id: Optional[int] = None,
        offset: int = 0,
        limit: int = 5000,
    ):
        """
        Multi-table join + JSON flatten + Feature Engineering → Master Dataset & RAG Tables.
        """
        # --- 1. Query with eager joins ---
        query = (
            db.query(Call)
            .join(Employee)
            .options(
                joinedload(Call.employee).joinedload(Employee.mastery_stats),
                joinedload(Call.campaign),
                joinedload(Call.outcome),
                joinedload(Call.qa_pairs),
                joinedload(Call.annotations),
            )
            .filter(Call.status.in_([CallStatus.EVALUATED, CallStatus.TRANSCRIBED]))
        )

        if campaign_id:
            query = query.filter(Call.campaign_id == campaign_id)
        if department:
            query = query.filter(Employee.department == department)
        if agent_role:
            query = query.filter(Employee.role == agent_role)
        if qa_scope_team_id is not None:
            query = query.join(
                EmployeeTeamAssignment,
                (Call.employee_id == EmployeeTeamAssignment.employee_id) & (EmployeeTeamAssignment.is_active == True),
            ).filter(EmployeeTeamAssignment.team_id == qa_scope_team_id)
        if qa_scope_campaign_id is not None:
            query = query.filter(Call.campaign_id == qa_scope_campaign_id)
        if start_date:
            query = query.filter(Call.created_at >= start_date)
        if end_date:
            query = query.filter(Call.created_at <= end_date)

        calls: List[Call] = query.order_by(Call.id).offset(offset).limit(limit).all()

        rows = []
        qa_rows = []
        annotation_rows = []

        for call in calls:
            emp = call.employee
            camp = call.campaign
            outcome = call.outcome
            mastery = emp.mastery_stats if emp else None

            # --- Temporal Features ---
            dt: Optional[datetime] = call.call_datetime or call.created_at
            call_hour = call.call_hour if call.call_hour is not None else (dt.hour if dt else None)
            call_day = call.call_day_of_week or (DAY_NAMES[dt.weekday()] if dt else "")
            call_date = dt.strftime("%Y-%m-%d %H:%M") if dt else ""

            # --- Acoustic Delta (Task 67) ---
            timeline = _safe_list(call.emotion_timeline)
            emotion_start = timeline[0].get("emotion", "neutral") if timeline else "neutral"
            emotion_end = timeline[-1].get("emotion", "neutral") if timeline else "neutral"
            
            # Use saved de_escalation_success if available, else fallback
            de_escalation = call.de_escalation_success
            if de_escalation is None:
                negative = ["stress", "agitation", "angry", "sad"]
                positive = ["calm", "happy", "neutral"]
                de_escalation = emotion_start.lower() in negative and emotion_end.lower() in positive

            # --- Campaign-Specific Flattening ---
            csd = _safe_json(outcome.campaign_specific_data if outcome else None)

            # --- Acoustic Emotion % ---
            emo_pcts = _calc_emotion_percentages(call.emotion_timeline)

            row = {
                # Identity
                "call_id": call.id,
                "call_date": call_date,
                "call_hour": call_hour,
                "call_day": call_day,
                "status": call.status.value if hasattr(call.status, "value") else str(call.status),
                # Agent
                "agent_id": emp.id if emp else None,
                "agent_name": emp.name if emp else "",
                "agent_code": emp.employee_code if emp else "",
                "agent_department": emp.department if emp else "",
                "agent_tier": (emp.tier.value if hasattr(emp.tier, "value") else str(emp.tier)) if emp else "",
                "agent_tenure_days": emp.agent_tenure_days if emp else None,
                # Campaign
                "campaign_id": camp.id if camp else None,
                "campaign_name": camp.name if camp else "",
                "campaign_type": (camp.type.value if hasattr(camp.type, "value") else str(camp.type)) if camp else "",
                "campaign_status": (camp.status.value if hasattr(camp.status, "value") else str(camp.status)) if camp else "",
                # Mastery
                "mastery_rapport": mastery.rapport_building if mastery else None,
                "mastery_sync": mastery.emotional_sync if mastery else None,
                "mastery_trust": mastery.ownership_trust if mastery else None,
                "mastery_clarity": mastery.process_clarity if mastery else None,
                # Evaluation
                "qa_score": call.evaluation_score,
                "overridden_score": call.overridden_score,
                "final_score": call.overridden_score if call.overridden_score is not None else call.evaluation_score,
                "is_golden_moment": call.is_golden_moment,
                "lead_status": call.lead_status.value if hasattr(call.lead_status, "value") else str(call.lead_status or ""),
                "opening_ok": call.opening_ok,
                "closing_ok": call.closing_ok,
                "dob_verified": call.dob_verified,
                # Talk
                "audio_duration_sec": call.audio_duration,
                "audio_duration_fmt": _seconds_to_mmss(call.audio_duration),
                "agent_talk_time_sec": call.agent_talk_time,
                "agent_talk_time_fmt": _seconds_to_mmss(call.agent_talk_time),
                "customer_talk_time_sec": call.customer_talk_time,
                "customer_talk_time_fmt": _seconds_to_mmss(call.customer_talk_time),
                "talk_ratio_pct": round((outcome.talk_ratio * 100), 1) if outcome and outcome.talk_ratio is not None else None,
                # Acoustic
                **emo_pcts,
                # Delta
                "emotion_start": emotion_start,
                "emotion_end": emotion_end,
                "de_escalation_success": de_escalation,
                # NLP
                "filler_words_count": call.filler_words_count,
                "interruptions_count": call.interruptions_count,
                "avg_response_time_sec": call.avg_response_time_sec,
                "calls_before_this": call.calls_before_this,
                # Outcome
                "primary_outcome": outcome.primary_outcome if outcome else "",
                "outcome_value": outcome.outcome_value if outcome else None,
                "follow_up_required": outcome.follow_up_required if outcome else False,
                "issue_resolved": csd.get("issue_resolved"),
                "resolution_type": csd.get("resolution_type", ""),
                "escalated": csd.get("escalated"),
                "customer_satisfaction": csd.get("customer_satisfaction"),
                "repeat_call": csd.get("repeat_call"),
            }
            rows.append(row)

            # RAG QA Pairs
            from app.routers.export import redact_text
            for pair in call.qa_pairs:
                objection = pair.objection
                response = pair.response
                if current_user_role != UserRole.ADMIN:
                    objection = redact_text(objection)
                    response = redact_text(response)
                qa_rows.append({
                    "call_id": call.id,
                    "objection": objection,
                    "response": response,
                    "emotion_at": pair.customer_emotion_at,
                    "emotion_after": pair.customer_emotion_after,
                    "is_golden": pair.is_golden_response
                })

            # Annotations
            for ann in call.annotations:
                note = ann.note
                if current_user_role != UserRole.ADMIN:
                    note = redact_text(note)
                annotation_rows.append({
                    "call_id": call.id,
                    "timestamp_sec": ann.timestamp,
                    "timestamp_fmt": _seconds_to_mmss(ann.timestamp),
                    "note": note,
                    "tag": ann.tag
                })

        df_master = pd.DataFrame(rows)
        if not df_master.empty:
            ordered = [c for c in cls.COLUMN_ORDER if c in df_master.columns]
            df_master = df_master[ordered]

        return df_master, pd.DataFrame(qa_rows), pd.DataFrame(annotation_rows)

    @classmethod
    def to_styled_xlsx(cls, df_master: pd.DataFrame, df_qa: pd.DataFrame, df_ann: pd.DataFrame) -> io.BytesIO:
        """
        Write multiple DataFrames to a styled .xlsx buffer (Multi-Sheet).
        """
        buffer = io.BytesIO()

        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            sheets = [
                ("Master Dataset", df_master),
                ("RAG QA Pairs", df_qa),
                ("Supervisor Annotations", df_ann)
            ]

            for sheet_name, df in sheets:
                if df.empty:
                    # Create empty sheet with headers if empty
                    pd.DataFrame(columns=df.columns).to_excel(writer, index=False, sheet_name=sheet_name)
                else:
                    df.to_excel(writer, index=False, sheet_name=sheet_name)
                
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]

                # --- Styling ---
                header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
                header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                thin_border = Border(
                    left=Side(style="thin", color="D1D5DB"),
                    right=Side(style="thin", color="D1D5DB"),
                    top=Side(style="thin", color="D1D5DB"),
                    bottom=Side(style="thin", color="D1D5DB"),
                )

                for col_idx, col_name in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                    
                    # Auto-width
                    max_len = max(len(str(col_name)), df[col_name].astype(str).str.len().max() if not df.empty else 0)
                    worksheet.column_dimensions[cell.column_letter].width = min(max_len + 4, 50)

                worksheet.freeze_panes = "A2"

        buffer.seek(0)
        return buffer
